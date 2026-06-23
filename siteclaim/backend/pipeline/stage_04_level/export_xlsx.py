"""Excel export of the levelled comparison — a professional tender bid-adjudication
workbook (openpyxl, imported lazily so the leveling math / DEMO_MODE never depend
on it).

Sheets:
  * Summary                  — title block + a ranking table per work section,
                               recommended tenderer per section.
  * <one per section>        — the Schedule-of-Rates comparison (Item · Desc · Unit ·
                               Qty · per-tenderer Rate/Amount · scheduled-rate
                               benchmark · variance · remarks), with corrected /
                               normalised subtotals and above-benchmark rates flagged.
  * Arithmetic Corrections   — every qty×rate correction (stated vs computed).
  * Scope Normalisation      — each scope gap added back at the peer rate, showing the
                               like-for-like flip.
  * Qualifications & Exclusions — each tenderer's stated exclusions / assumptions.

All values are read from the levelled bids, the bid line items (Qty/Unit), the
arithmetic findings, the scope gaps, and the "tender-scheduled-rates" benchmark bid
already present in the replies — nothing is invented here.
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Optional

from rules_engine.leveling import peer_item_reference
from schemas.models import BidReply, LevelledBid

OUT_PATH = Path(__file__).resolve().parents[2] / "fixtures" / "out" / "leveling.xlsx"

BENCHMARK_ID = "tender-scheduled-rates"
_CUR = '"HK$"#,##0'
_PCT = "+0.0%;-0.0%;0.0%"

_TITLES = {
    "field_testing": "Field Testing",
    "field_installations": "Field Installations",
    "geophysical_survey": "Geophysical Survey",
    "drilling": "Drilling",
    "sampling": "Sampling",
    "electrical": "Electrical",
    "mechanical_plumbing": "Mechanical & Plumbing",
    "fire_services": "Fire Services",
    "joinery_fitting_out": "Joinery & Fitting-out",
}


def _title(trade: str) -> str:
    return _TITLES.get(trade, trade.replace("_", " ").title())


def _sheet_name(name: str, used: set[str]) -> str:
    base = name[:31]
    candidate, i = base, 1
    while candidate in used:
        i += 1
        candidate = f"{base[:28]}~{i}"
    used.add(candidate)
    return candidate


def export_leveling_xlsx(
    levelled: list[LevelledBid],
    replies: list[BidReply],
    item_order: Optional[list[str]] = None,
    path: Path | str = OUT_PATH,
    project_name: str = "",
) -> Path:
    """Write the adjudication workbook to ``path`` and return it."""
    from openpyxl import Workbook  # lazy — leveling math must not require openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # ---- shared styles -----------------------------------------------------
    INK = "0F1B2D"
    band = PatternFill("solid", fgColor=INK)
    band_font = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=15, color=INK)
    sub_font = Font(size=9.5, italic=True, color="6B7A90")
    bold = Font(bold=True)
    subtotal_fill = PatternFill("solid", fgColor="EEF2F8")
    flag_fill = PatternFill("solid", fgColor="FCE4E4")  # rate above benchmark
    rec_fill = PatternFill("solid", fgColor="E5F4EC")   # recommended row
    thin = Side(style="thin", color="D5DCE6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    centre = Alignment(horizontal="center")

    # ---- data indexes ------------------------------------------------------
    trades: list[str] = []
    for b in levelled:
        if b.trade not in trades:
            trades.append(b.trade)
    line_of = {(r.firm_id, r.trade, ln.item_ref): ln for r in replies for ln in r.line_items}
    reply_of = {(r.firm_id, r.trade): r for r in replies}
    peer = peer_item_reference(replies)

    def section_bids(trade: str):
        bids = [b for b in levelled if b.trade == trade]
        tenderers = [b for b in bids if b.firm_id != BENCHMARK_ID]
        benchmark = next((b for b in bids if b.firm_id == BENCHMARK_ID), None)
        return tenderers, benchmark

    def ranked(tenderers: list[LevelledBid]) -> list[LevelledBid]:
        return sorted(tenderers, key=lambda b: (b.normalized_total, b.corrected_total))

    def section_items(trade: str, tenderers, benchmark) -> list[str]:
        order: list[str] = []
        for b in [*tenderers, *( [benchmark] if benchmark else [])]:
            rep = reply_of.get((b.firm_id, trade))
            for ln in (rep.line_items if rep else []):
                if ln.item_ref not in order:
                    order.append(ln.item_ref)
        if item_order:  # honour an explicit hint where it overlaps this section
            hinted = [i for i in item_order if i in order]
            order = hinted + [i for i in order if i not in hinted]
        return order

    def fit(ws, widths: dict[int, int]) -> None:
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

    def band_row(ws, row: int, ncols: int) -> None:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill, cell.font, cell.border = band, band_font, border

    wb = Workbook()
    used_names: set[str] = set()

    # =======================================================================
    # SHEET: Summary
    # =======================================================================
    ws = wb.active
    ws.title = _sheet_name("Summary", used_names)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "Tender Bid Adjudication — Levelled Comparison"
    ws["A1"].font = title_font
    meta = [
        ("Project", project_name or "Tender — levelled bid comparison"),
        ("Employer", "________________________"),
        ("Main Contractor", "________________________"),
        ("Date", _dt.date.today().strftime("%d %b %Y")),
        ("Prepared by", "________________________"),
        ("Checked by", "________________________"),
    ]
    r = 2
    for k, v in meta:
        ws.cell(row=r, column=1, value=k).font = bold
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=v)
        r += 1
    ws.cell(row=r, column=1, value="Commercial in confidence").font = sub_font
    r += 2

    rec_lines: list[tuple[str, str]] = []
    for trade in trades:
        tenderers, benchmark = section_bids(trade)
        order = ranked(tenderers)
        rec = order[0] if order else None
        if rec:
            rec_lines.append((_title(trade), rec.firm_name))

        ws.cell(row=r, column=1, value=f"Section — {_title(trade)}").font = Font(bold=True, size=11, color=INK)
        r += 1
        head = ["Tenderer", "Corrected sum", "Normalised sum", "Rank", "Recommended", "Basis"]
        for c, h in enumerate(head, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill, cell.font, cell.border = band, band_font, border
            cell.alignment = centre if c in (4, 5) else left
        r += 1
        cheapest_corrected = min((b.corrected_total for b in tenderers), default=0.0)
        for i, b in enumerate(order, start=1):
            recommended = i == 1
            flip = recommended and b.corrected_total > cheapest_corrected + 0.5
            basis = (
                ("Lowest normalised tender sum after scope normalisation (a lower-priced bid omitted scope)"
                 if flip else "Lowest normalised (like-for-like) tender sum")
                if recommended else "Higher normalised tender sum"
            )
            vals = [b.firm_name, b.corrected_total, b.normalized_total, i, "Y" if recommended else "N", basis]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = border
                if c in (2, 3):
                    cell.number_format, cell.alignment = _CUR, right
                elif c in (4, 5):
                    cell.alignment = centre
                else:
                    cell.alignment = left
                if recommended:
                    cell.fill = rec_fill
                    if c in (1, 5):
                        cell.font = bold
            r += 1
        if benchmark:
            vals = [f"{benchmark.firm_name}", benchmark.corrected_total, benchmark.normalized_total, "—", "—", "Tender Schedule of Rates — baseline, not a tenderer"]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border, cell.font = border, sub_font
                if c in (2, 3):
                    cell.number_format, cell.alignment = _CUR, right
                elif c in (4, 5):
                    cell.alignment = centre
                else:
                    cell.alignment = left
            r += 1
        r += 1

    ws.cell(row=r, column=1, value="Recommended award by section").font = bold
    r += 1
    for sect, firm in rec_lines:
        ws.cell(row=r, column=1, value=sect).font = bold
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=firm)
        r += 1
    fit(ws, {1: 34, 2: 16, 3: 16, 4: 7, 5: 13, 6: 52})
    ws.freeze_panes = "A2"

    # =======================================================================
    # SHEET per section: the SoR comparison
    # =======================================================================
    finding_items = {(b.firm_id, b.trade): {f.location.replace("line ", "") for f in b.arithmetic_findings} for b in levelled}
    gap_items = {(b.firm_id, b.trade): {g.split(" — ")[0] for g in b.scope_gaps} for b in levelled}

    for trade in trades:
        tenderers, benchmark = section_bids(trade)
        order = section_items(trade, tenderers, benchmark)
        wss = wb.create_sheet(_sheet_name(_title(trade), used_names))
        wss.sheet_view.showGridLines = False

        ncols = 4 + 2 * len(tenderers) + (2 if benchmark else 0) + 1
        wss.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        wss.cell(row=1, column=1, value=f"{_title(trade)} — Schedule of Rates comparison (levelled)").font = title_font

        # header row (row 2)
        head = ["Item", "Description", "Unit", "Qty"]
        rate_col, amt_col = {}, {}
        col = 5
        for b in tenderers:
            rate_col[b.firm_id], amt_col[b.firm_id] = col, col + 1
            head += [f"{b.firm_name} — Rate", f"{b.firm_name} — Amount"]
            col += 2
        sched_col = var_col = None
        if benchmark:
            sched_col, var_col = col, col + 1
            head += ["Scheduled Rate", "Variance vs benchmark %"]
            col += 2
        remarks_col = col
        head.append("Remarks / flags")
        for c, h in enumerate(head, start=1):
            wss.cell(row=2, column=c, value=h)
        band_row(wss, 2, ncols)
        for c in range(1, ncols + 1):
            wss.cell(row=2, column=c).alignment = Alignment(horizontal="center", wrap_text=True)

        row = 3
        for item in order:
            sample = next((line_of.get((b.firm_id, trade, item)) for b in [*tenderers, *( [benchmark] if benchmark else [])] if line_of.get((b.firm_id, trade, item))), None)
            if sample is None:
                continue
            provisional = "provisional" in (sample.description or "").lower()
            wss.cell(row=row, column=1, value=item)
            wss.cell(row=row, column=2, value=sample.description)
            wss.cell(row=row, column=3, value=sample.unit)
            qcell = wss.cell(row=row, column=4, value=sample.qty)
            qcell.number_format = "#,##0.##"
            remarks: list[str] = []
            bench_line = line_of.get((BENCHMARK_ID, trade, item)) if benchmark else None
            bench_rate = bench_line.rate if bench_line else None
            for b in tenderers:
                ln = line_of.get((b.firm_id, trade, item))
                rc = wss.cell(row=row, column=rate_col[b.firm_id])
                ac = wss.cell(row=row, column=amt_col[b.firm_id])
                if ln is None or ln.rate is None:
                    rc.value, ac.value = "—", "—"
                    rc.alignment = ac.alignment = right
                    if item in gap_items.get((b.firm_id, trade), set()):
                        remarks.append(f"{b.firm_name}: not priced — scope gap")
                else:
                    rc.value, rc.number_format, rc.alignment = ln.rate, _CUR, right
                    ac.value, ac.number_format, ac.alignment = round(ln.qty * ln.rate, 2), _CUR, right
                    if bench_rate not in (None, 0) and ln.rate > bench_rate:
                        rc.fill = flag_fill
                        remarks.append(f"{b.firm_name}: rate above benchmark")
                if item in finding_items.get((b.firm_id, trade), set()):
                    remarks.append(f"{b.firm_name}: arithmetic corrected")
            if benchmark:
                sc = wss.cell(row=row, column=sched_col)
                if bench_rate is not None:
                    sc.value, sc.number_format, sc.alignment = bench_rate, _CUR, right
                else:
                    sc.value, sc.alignment = "—", right
                vc = wss.cell(row=row, column=var_col)
                lead = line_of.get((tenderers[0].firm_id, trade, item)) if tenderers else None
                if lead and lead.rate is not None and bench_rate not in (None, 0):
                    vc.value = (lead.rate - bench_rate) / bench_rate
                    vc.number_format, vc.alignment = _PCT, right
                else:
                    vc.value, vc.alignment = "—", right
            if provisional:
                remarks.append("Provisional sum — carried separately")
            rcell = wss.cell(row=row, column=remarks_col, value="; ".join(dict.fromkeys(remarks)))
            rcell.alignment = left
            for c in range(1, ncols + 1):
                wss.cell(row=row, column=c).border = border
            row += 1

        # subtotal rows: corrected then normalised
        for label, attr in (("Corrected tender sum", "corrected_total"), ("Normalised tender sum (like-for-like)", "normalized_total")):
            wss.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            lc = wss.cell(row=row, column=1, value=label)
            lc.font = bold
            for b in tenderers:
                cell = wss.cell(row=row, column=amt_col[b.firm_id], value=getattr(b, attr))
                cell.number_format, cell.font, cell.alignment = _CUR, bold, right
            if benchmark:
                cell = wss.cell(row=row, column=sched_col, value=getattr(benchmark, attr))
                cell.number_format, cell.font, cell.alignment = _CUR, bold, right
            for c in range(1, ncols + 1):
                cell = wss.cell(row=row, column=c)
                cell.border, cell.fill = border, subtotal_fill
                if not cell.font.bold:
                    cell.font = bold
            row += 1

        widths = {1: 9, 2: 44, 3: 9, 4: 8}
        for b in tenderers:
            widths[rate_col[b.firm_id]] = 13
            widths[amt_col[b.firm_id]] = 14
        if benchmark:
            widths[sched_col] = 14
            widths[var_col] = 13
        widths[remarks_col] = 40
        fit(wss, widths)
        wss.freeze_panes = "A3"

    # =======================================================================
    # SHEET: Arithmetic Corrections
    # =======================================================================
    wsa = wb.create_sheet(_sheet_name("Arithmetic Corrections", used_names))
    wsa.sheet_view.showGridLines = False
    head = ["Tenderer", "Section", "Item", "Stated amount", "Computed (Qty × Rate)", "Corrected"]
    for c, h in enumerate(head, start=1):
        wsa.cell(row=1, column=c, value=h)
    band_row(wsa, 1, len(head))
    row = 2
    for b in levelled:
        for f in b.arithmetic_findings:
            item = f.location.replace("line ", "")
            ln = line_of.get((b.firm_id, b.trade, item))
            stated = ln.amount if ln else None
            vals = [b.firm_name, _title(b.trade), item, stated, f.corrected_value, f.corrected_value]
            for c, v in enumerate(vals, start=1):
                cell = wsa.cell(row=row, column=c, value=v)
                cell.border = border
                if c >= 4:
                    cell.number_format, cell.alignment = _CUR, right
            row += 1
    if row == 2:
        wsa.merge_cells("A2:F2")
        wsa.cell(row=2, column=1, value="No arithmetic corrections — all stated amounts tie to Qty × Rate.").font = sub_font
    fit(wsa, {1: 34, 2: 18, 3: 9, 4: 16, 5: 20, 6: 16})
    wsa.freeze_panes = "A2"

    # =======================================================================
    # SHEET: Scope Normalisation
    # =======================================================================
    wsn = wb.create_sheet(_sheet_name("Scope Normalisation", used_names))
    wsn.sheet_view.showGridLines = False
    wsn.merge_cells("A1:F1")
    wsn.cell(row=1, column=1, value="Scope normalisation — unpriced items added back at the peer (median) rate so every bid is compared like-for-like").font = sub_font
    head = ["Tenderer", "Section", "Item / description", "Peer amount added", "Corrected sum", "Normalised sum"]
    for c, h in enumerate(head, start=1):
        wsn.cell(row=2, column=c, value=h)
    band_row(wsn, 2, len(head))
    row = 3
    any_gap = False
    for b in levelled:
        if not b.scope_gaps:
            continue
        any_gap = True
        first = True
        for gap in b.scope_gaps:
            item = gap.split(" — ")[0]
            added = peer.get(item, 0.0)
            desc = gap.split(" — ", 1)[1] if " — " in gap else gap
            vals = [b.firm_name if first else "", _title(b.trade) if first else "", f"{item} — {desc}", added,
                    b.corrected_total if first else "", b.normalized_total if first else ""]
            for c, v in enumerate(vals, start=1):
                cell = wsn.cell(row=row, column=c, value=v)
                cell.border = border
                if c in (4, 5, 6) and isinstance(v, (int, float)):
                    cell.number_format, cell.alignment = _CUR, right
                if c == 3:
                    cell.alignment = left
            first = False
            row += 1
    if not any_gap:
        wsn.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        wsn.cell(row=row, column=1, value="No scope gaps — every tenderer priced the full scope.").font = sub_font
    fit(wsn, {1: 34, 2: 18, 3: 46, 4: 18, 5: 16, 6: 16})
    wsn.freeze_panes = "A3"

    # =======================================================================
    # SHEET: Qualifications & Exclusions
    # =======================================================================
    wsq = wb.create_sheet(_sheet_name("Qualifications & Exclusions", used_names))
    wsq.sheet_view.showGridLines = False
    wsq.merge_cells("A1:C1")
    wsq.cell(row=1, column=1, value="Stated exclusions and assumptions are flagged as non-comparable and are NOT deducted from the tender sum.").font = sub_font
    head = ["Tenderer", "Section", "Stated exclusion / assumption"]
    for c, h in enumerate(head, start=1):
        wsq.cell(row=2, column=c, value=h)
    band_row(wsq, 2, len(head))
    row = 3
    any_excl = False
    for b in levelled:
        for ex in b.exclusions:
            any_excl = True
            wsq.cell(row=row, column=1, value=b.firm_name).border = border
            wsq.cell(row=row, column=2, value=_title(b.trade)).border = border
            cell = wsq.cell(row=row, column=3, value=ex)
            cell.border, cell.alignment = border, left
            row += 1
    if not any_excl:
        wsq.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        wsq.cell(row=row, column=1, value="No stated exclusions or qualifications.").font = sub_font
    fit(wsq, {1: 34, 2: 18, 3: 70})
    wsq.freeze_panes = "A3"

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
